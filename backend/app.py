"""MES工厂管家 - 后端主程序（重构版）"""
import os
import sys
import json
import csv
import io
import datetime
import secrets
import sqlite3

# 确保可以导入模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, request, jsonify, send_from_directory, make_response, session
from openpyxl import Workbook, load_workbook

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    _HAS_LIMITER = True
except ImportError:  # pragma: no cover
    _HAS_LIMITER = False

from utils.rate_limiter import SimpleRateLimiter

from utils.database import close_db, init_db, _init_extra_tables, DB_PATH, BASE_DIR, get_db
from utils.db_errors import INTEGRITY_ERRORS
from utils.helpers import login_required, gen_no, _load_session_user
from blueprints.auth import auth_bp
from blueprints.system import system_bp
from blueprints.base_data import base_data_bp
from blueprints.inventory import inventory_bp
from blueprints.production import production_bp
from blueprints.quality import quality_bp
from blueprints.equipment import equipment_bp
from blueprints.tool import tool_bp
from blueprints.schedule import schedule_bp
from blueprints.flow import flow_bp
from blueprints.dashboard import dashboard_bp
from blueprints.report import report_bp
from blueprints.notification import notification_bp
from blueprints.supplier import supplier_bp
from blueprints.customer import customer_bp
from blueprints.document import document_bp
from blueprints.cost import cost_bp
from blueprints.barcode import barcode_bp
from blueprints.backup import backup_bp
from blueprints.kanban import kanban_bp
from blueprints.security import security_bp
from blueprints.prod_ext import prod_ext_bp
from blueprints.qm_ext import qm_ext_bp
from blueprints.analytics import analytics_bp
from blueprints.sys_ext import sys_ext_bp
from blueprints.site import site_bp
from blueprints.qm_plus import qm_plus_bp
from blueprints.eqp_plus import eqp_plus_bp
from blueprints.util import util_bp
from blueprints.hr import hr_bp
from blueprints.five_s import five_s_bp
from blueprints.svc import svc_bp
from blueprints.search import search_bp
from blueprints.stage import stage_bp
from blueprints.process_ctrl import process_bp
from blueprints.erp import erp_bp
from blueprints.iot import iot_bp
from blueprints.ai import ai_bp
from blueprints.scm import scm_bp
from blueprints.aps import aps_bp
from blueprints.fmea import fmea_bp
from blueprints.tenant import tenant_bp
from blueprints.erp_deep import erp_deep_bp
from blueprints.update import update_bp
from blueprints.sop_warehouse import sop_bp
from blueprints.trace_ext import trace_ext_bp
from blueprints.warehouse import warehouse_bp
from blueprints.eqp_schedule import eqp_schedule_bp
from blueprints.table_order import table_order_bp
from blueprints.machine_iot import machine_iot_bp
from blueprints.device_platform import device_platform_bp

FRONTEND_DIR = os.path.join(BASE_DIR, 'frontend')
ADMIN_DIR = os.path.join(BASE_DIR, 'admin')


def create_app():
    app = Flask(__name__, static_folder=None)
    secret_key = os.environ.get('FLASK_SECRET_KEY') or os.environ.get('SECRET_KEY')
    if not secret_key and os.environ.get('MES_ENV', '').lower() == 'production':
        raise RuntimeError('FLASK_SECRET_KEY is required in production')
    app.secret_key = secret_key or secrets.token_hex(32)
    app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MES_MAX_REQUEST_BYTES', 512 * 1024))
    app.config['SESSION_COOKIE_NAME'] = os.environ.get(
        'SESSION_COOKIE_NAME', 'mes_main_session'
    )
    # Session cookie security: always HttpOnly, Secure only in production, SameSite=Lax
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SECURE'] = (
        os.environ.get('MES_ENV', '').lower() == 'production'
    )
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

    # 限流:优先使用 flask-limiter;未安装时使用内置内存限流兜底,避免生产环境无保护
    if _HAS_LIMITER:
        Limiter(
            key_func=get_remote_address,
            default_limits=['200 per day', '50 per hour'],
            storage_uri='memory://',
        ).init_app(app)
    else:
        SimpleRateLimiter(
            default_limits=['200 per day', '50 per hour'],
        ).init_app(app)

    # 注册 teardown
    app.teardown_appcontext(close_db)

    # 注册蓝图
    app.register_blueprint(auth_bp)
    app.register_blueprint(system_bp)
    app.register_blueprint(base_data_bp)
    app.register_blueprint(inventory_bp)
    app.register_blueprint(production_bp)
    app.register_blueprint(quality_bp)
    app.register_blueprint(equipment_bp)
    app.register_blueprint(tool_bp)
    app.register_blueprint(schedule_bp)
    app.register_blueprint(flow_bp)
    app.register_blueprint(dashboard_bp)
    app.register_blueprint(report_bp)
    app.register_blueprint(notification_bp)
    app.register_blueprint(supplier_bp)
    app.register_blueprint(customer_bp)
    app.register_blueprint(document_bp)
    app.register_blueprint(cost_bp)
    app.register_blueprint(barcode_bp)
    app.register_blueprint(backup_bp)
    app.register_blueprint(kanban_bp)
    app.register_blueprint(security_bp)
    app.register_blueprint(prod_ext_bp)
    app.register_blueprint(qm_ext_bp)
    app.register_blueprint(analytics_bp)
    app.register_blueprint(sys_ext_bp)
    app.register_blueprint(site_bp)
    app.register_blueprint(qm_plus_bp)
    app.register_blueprint(eqp_plus_bp)
    app.register_blueprint(util_bp)
    app.register_blueprint(hr_bp)
    app.register_blueprint(five_s_bp)
    app.register_blueprint(svc_bp)
    app.register_blueprint(search_bp)
    app.register_blueprint(stage_bp)
    app.register_blueprint(process_bp)
    app.register_blueprint(erp_bp)
    app.register_blueprint(iot_bp)
    app.register_blueprint(ai_bp)
    app.register_blueprint(scm_bp)
    app.register_blueprint(aps_bp)
    app.register_blueprint(fmea_bp)
    app.register_blueprint(tenant_bp)
    app.register_blueprint(erp_deep_bp)
    app.register_blueprint(update_bp)
    app.register_blueprint(sop_bp)
    app.register_blueprint(trace_ext_bp)
    app.register_blueprint(warehouse_bp)
    app.register_blueprint(eqp_schedule_bp)
    app.register_blueprint(table_order_bp)
    app.register_blueprint(machine_iot_bp)
    app.register_blueprint(device_platform_bp)

    # 静态文件路由
    @app.route('/')
    def index():
        return send_from_directory(FRONTEND_DIR, 'index.html')

    # 健康检查路由(由 utils.health_checks 提供)
    from utils.health_checks import register_health_routes
    register_health_routes(app)

    @app.route('/manifest.json')
    def manifest():
        return send_from_directory(FRONTEND_DIR, 'manifest.json')

    @app.route('/frontend/static/<path:filename>')
    def frontend_static(filename):
        return send_from_directory(os.path.join(FRONTEND_DIR, 'static'), filename)

    @app.route('/admin')
    def admin_page():
        return send_from_directory(ADMIN_DIR, 'index.html')

    @app.route('/admin/static/<path:filename>')
    def admin_static(filename):
        return send_from_directory(os.path.join(ADMIN_DIR, 'static'), filename)

    # 导入导出功能
    TABLE_CONFIG = {
        'base_product': {
            'name': '产品',
            'columns': ['product_name', 'code', 'specification', 'unit', 'product_type', 'status'],
            'headers': ['产品名称', '产品编码', '规格型号', '单位', '类型', '状态'],
            'types': ['str', 'str', 'str', 'str', 'str', 'int']
        },
        'base_process': {
            'name': '工序',
            'columns': ['process_name', 'code', 'workshop_id', 'description', 'standard_time', 'status'],
            'headers': ['工序名称', '工序编码', '车间ID', '描述', '标准工时(分钟)', '状态'],
            'types': ['str', 'str', 'int', 'str', 'float', 'int']
        },
        'base_workshop': {
            'name': '车间',
            'columns': ['workshop_name', 'code', 'description', 'status'],
            'headers': ['车间名称', '车间编码', '描述', '状态'],
            'types': ['str', 'str', 'str', 'int']
        },
        'prod_workorder': {
            'name': '工单',
            'columns': ['order_no', 'product_id', 'planned_qty', 'priority', 'status', 'remark'],
            'headers': ['工单号', '产品ID', '计划数量', '优先级', '状态', '备注'],
            'types': ['str', 'int', 'float', 'int', 'int', 'str']
        },
        'prod_task': {
            'name': '任务',
            'columns': ['task_no', 'workorder_id', 'process_id', 'planned_qty', 'status', 'remark'],
            'headers': ['任务编号', '工单ID', '工序ID', '计划数量', '状态', '备注'],
            'types': ['str', 'int', 'int', 'float', 'int', 'str']
        },
        'prod_report': {
            'name': '报工',
            'columns': ['report_no', 'task_id', 'workorder_id', 'process_id', 'qualified_qty', 'defect_qty', 'remark'],
            'headers': ['报工单号', '任务ID', '工单ID', '工序ID', '合格数量', '不良数量', '备注'],
            'types': ['str', 'int', 'int', 'int', 'float', 'float', 'str']
        },
        'inv_inbound': {
            'name': '入库单',
            'columns': ['inbound_no', 'inbound_type', 'supplier', 'total_amount', 'status', 'remark'],
            'headers': ['入库单号', '类型', '供应商', '金额', '状态', '备注'],
            'types': ['str', 'str', 'str', 'float', 'int', 'str']
        },
        'inv_outbound': {
            'name': '出库单',
            'columns': ['outbound_no', 'outbound_type', 'customer', 'total_amount', 'status', 'remark'],
            'headers': ['出库单号', '类型', '客户', '金额', '状态', '备注'],
            'types': ['str', 'str', 'str', 'float', 'int', 'str']
        },
        'inv_balance': {
            'name': '库存',
            'columns': ['product_id', 'quantity', 'amount'],
            'headers': ['产品ID', '数量', '金额'],
            'types': ['int', 'float', 'float']
        },
        'tool_ledger': {
            'name': '工具',
            'columns': ['tool_name', 'code', 'type_id', 'specification', 'quantity', 'location', 'status'],
            'headers': ['工具名称', '工具编码', '类型ID', '规格', '数量', '位置', '状态'],
            'types': ['str', 'str', 'int', 'str', 'float', 'str', 'int']
        },
        'sys_user': {
            'name': '用户',
            'columns': ['username', 'real_name', 'phone', 'email', 'dept_id', 'role_id', 'status'],
            'headers': ['用户名', '姓名', '电话', '邮箱', '部门ID', '角色ID', '状态'],
            'types': ['str', 'str', 'str', 'str', 'int', 'int', 'int']
        },
        'sys_dept': {
            'name': '部门',
            'columns': ['dept_name', 'leader', 'phone', 'sort_order', 'status'],
            'headers': ['部门名称', '负责人', '电话', '排序', '状态'],
            'types': ['str', 'str', 'str', 'int', 'int']
        },
        'base_bom': {
            'name': '物料清单',
            'columns': ['product_id', 'material_id', 'quantity', 'unit', 'description'],
            'headers': ['产品ID', '物料ID', '用量', '单位', '描述'],
            'types': ['int', 'int', 'float', 'str', 'str']
        },
        'base_defect': {
            'name': '不良品项',
            'columns': ['defect_name', 'code', 'defect_type', 'description', 'status'],
            'headers': ['缺陷名称', '编码', '类型', '描述', '状态'],
            'types': ['str', 'str', 'str', 'str', 'int']
        },
        'base_unit': {
            'name': '单位',
            'columns': ['unit_name', 'unit_symbol', 'status'],
            'headers': ['单位名称', '符号', '状态'],
            'types': ['str', 'str', 'int']
        },
        'base_process_route': {
            'name': '工艺路线',
            'columns': ['product_id', 'route_name', 'description', 'status'],
            'headers': ['产品ID', '路线名称', '描述', '状态'],
            'types': ['int', 'str', 'str', 'int']
        },
        'prod_sales_order': {
            'name': '销售订单',
            'columns': ['order_no', 'customer', 'contact', 'phone', 'total_amount', 'delivery_date', 'status', 'remark'],
            'headers': ['订单号', '客户', '联系人', '电话', '金额', '交货日期', '状态', '备注'],
            'types': ['str', 'str', 'str', 'str', 'float', 'str', 'int', 'str']
        },
        'prod_plan': {
            'name': '生产计划',
            'columns': ['plan_no', 'plan_type', 'start_date', 'end_date', 'status', 'remark'],
            'headers': ['计划编号', '类型', '开始日期', '结束日期', '状态', '备注'],
            'types': ['str', 'str', 'str', 'str', 'int', 'str']
        },
        'qm_incoming_inspection': {
            'name': '来料检验',
            'columns': ['inspect_no', 'supplier', 'result', 'status', 'remark'],
            'headers': ['检验单号', '供应商', '结果', '状态', '备注'],
            'types': ['str', 'str', 'str', 'int', 'str']
        },
        'qm_process_inspection': {
            'name': '过程检验',
            'columns': ['inspect_no', 'workorder_id', 'result', 'status', 'remark'],
            'headers': ['检验单号', '工单ID', '结果', '状态', '备注'],
            'types': ['str', 'int', 'str', 'int', 'str']
        },
        'qm_outgoing_inspection': {
            'name': '出货检验',
            'columns': ['inspect_no', 'customer', 'result', 'status', 'remark'],
            'headers': ['检验单号', '客户', '结果', '状态', '备注'],
            'types': ['str', 'str', 'str', 'int', 'str']
        },
        'sched_team': {
            'name': '班组',
            'columns': ['team_name', 'code', 'leader', 'member_count', 'workshop_id', 'status'],
            'headers': ['班组名称', '编码', '班组长', '人数', '车间ID', '状态'],
            'types': ['str', 'str', 'str', 'int', 'int', 'int']
        },
        'sched_plan': {
            'name': '排班计划',
            'columns': ['plan_name', 'team_id', 'start_date', 'end_date', 'shift_type', 'status'],
            'headers': ['计划名称', '班组ID', '开始日期', '结束日期', '班次', '状态'],
            'types': ['str', 'int', 'str', 'str', 'str', 'int']
        },
        'eqp_ledger': {
            'name': '设备台账',
            'columns': ['equipment_name', 'code', 'type_id', 'model', 'manufacturer', 'workshop_id', 'location', 'status', 'remark'],
            'headers': ['设备名称', '编码', '类型ID', '型号', '制造商', '车间ID', '位置', '状态', '备注'],
            'types': ['str', 'str', 'int', 'str', 'str', 'int', 'str', 'int', 'str']
        },
        'eqp_repair_order': {
            'name': '维修单',
            'columns': ['repair_no', 'equipment_id', 'fault_desc', 'repair_desc', 'status', 'remark'],
            'headers': ['维修单号', '设备ID', '故障描述', '维修描述', '状态', '备注'],
            'types': ['str', 'int', 'str', 'str', 'int', 'str']
        },
        'tool_type': {
            'name': '工具类型',
            'columns': ['type_name', 'code', 'description', 'status'],
            'headers': ['类型名称', '编码', '描述', '状态'],
            'types': ['str', 'str', 'str', 'int']
        },
        'tool_borrow': {
            'name': '工具领用',
            'columns': ['borrow_no', 'tool_id', 'borrower', 'borrow_qty', 'return_qty', 'status', 'remark'],
            'headers': ['领用单号', '工具ID', '领用人ID', '领用数量', '归还数量', '状态', '备注'],
            'types': ['str', 'int', 'int', 'float', 'float', 'int', 'str']
        },
        'sys_role': {
            'name': '角色',
            'columns': ['role_name', 'role_key', 'description', 'status'],
            'headers': ['角色名称', '标识', '描述', '状态'],
            'types': ['str', 'str', 'str', 'int']
        },
        'sys_dict': {
            'name': '数据字典',
            'columns': ['dict_type', 'dict_label', 'dict_value', 'sort_order', 'status'],
            'headers': ['类型', '标签', '值', '排序', '状态'],
            'types': ['str', 'str', 'str', 'int', 'int']
        },
        'inv_warehouse': {
            'name': '仓库',
            'columns': ['warehouse_name', 'code', 'address', 'status'],
            'headers': ['仓库名称', '仓库编码', '地址', '状态'],
            'types': ['str', 'str', 'str', 'int']
        },
        'inv_area': {
            'name': '库区',
            'columns': ['warehouse_id', 'area_name', 'code', 'status'],
            'headers': ['仓库ID', '库区名称', '库区编码', '状态'],
            'types': ['int', 'str', 'str', 'int']
        },
        'inv_location': {
            'name': '库位',
            'columns': ['area_id', 'location_name', 'code', 'status'],
            'headers': ['库区ID', '库位名称', '库位编码', '状态'],
            'types': ['int', 'str', 'str', 'int']
        },
        'inv_arrival_notice': {
            'name': '到货通知',
            'columns': ['notice_no', 'supplier_id', 'expected_date', 'status', 'remark'],
            'headers': ['通知单号', '供应商ID', '预计到货日', '状态', '备注'],
            'types': ['str', 'int', 'str', 'int', 'str']
        },
        'inv_transaction_log': {
            'name': '库存事务',
            'columns': [
                'trans_type', 'product_id', 'quantity', 'warehouse_id',
                'area_id', 'location_id', 'batch_no', 'ref_no',
                'ref_type', 'remark'
            ],
            'headers': [
                '事务类型', '产品ID', '数量', '仓库ID', '库区ID',
                '库位ID', '批次号', '关联单号', '关联类型', '备注'
            ],
            'types': [
                'str', 'int', 'float', 'int', 'int', 'int',
                'str', 'str', 'str', 'str'
            ],
            'importable': False
        },
        'qm_inspect_template': {
            'name': '质检模板',
            'columns': ['template_name', 'inspect_type', 'items', 'status'],
            'headers': ['模板名称', '检验类型', '检验项目JSON', '状态'],
            'types': ['str', 'str', 'str', 'int']
        },
        'eqp_check_project': {
            'name': '设备点检项目',
            'columns': ['project_name', 'check_type', 'standard', 'method', 'status'],
            'headers': ['项目名称', '点检类型', '点检标准', '点检方法', '状态'],
            'types': ['str', 'str', 'str', 'str', 'int']
        },
        'sched_calendar': {
            'name': '排班日历',
            'columns': ['plan_id', 'work_date', 'shift_type', 'user_ids'],
            'headers': ['排班计划ID', '工作日期', '班次类型', '人员ID'],
            'types': ['int', 'str', 'str', 'str']
        }
    }

    @app.route('/api/export/<table>')
    @login_required
    def export_data(table):
        if table not in TABLE_CONFIG:
            return jsonify({'code': 400, 'message': '不支持的表'})
        config = TABLE_CONFIG[table]
        db = get_db()
        rows = db.execute(f"SELECT * FROM {table}").fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = config['name']
        ws.append(config['headers'])

        for row in rows:
            row_data = []
            for col in config['columns']:
                row_data.append(row[col] if col in row.keys() else '')
            ws.append(row_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{table}_{timestamp}.xlsx"
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    @app.route('/api/import/<table>', methods=['POST'])
    @login_required
    def import_data(table):
        current_user, auth_error = _load_session_user()
        if auth_error:
            return auth_error
        if table.startswith('sys_') and (current_user['role_key'] != 'admin' or not current_user['role_status']):
            return jsonify({
                'code': 403,
                'message': '仅管理员可导入系统管理数据',
            }), 403
        if table not in TABLE_CONFIG:
            return jsonify({'code': 400, 'message': '不支持的表'})
        if TABLE_CONFIG[table].get('importable') is False:
            return jsonify({'code': 400, 'message': '不支持导入该表'})

        if 'file' not in request.files:
            return jsonify({'code': 400, 'message': '请选择文件'})

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'code': 400, 'message': '请上传Excel或CSV文件'})

        config = TABLE_CONFIG[table]
        db = get_db()
        success_count = 0
        error_rows = []

        try:
            if file.filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.reader(io.StringIO(content))
                headers = next(reader)
                for i, row in enumerate(reader, start=2):
                    try:
                        data = {}
                        for j, col in enumerate(config['columns']):
                            if j < len(row):
                                val = row[j].strip()
                                if config['types'][j] == 'int':
                                    data[col] = int(val) if val else None
                                elif config['types'][j] == 'float':
                                    data[col] = float(val) if val else 0
                                else:
                                    data[col] = val
                        if 'status' not in data:
                            data['status'] = 1
                        db.execute(f"INSERT INTO {table} ({','.join(config['columns'])}) VALUES ({','.join(['?']*len(config['columns']))})",
                                   [data.get(c) for c in config['columns']])
                        success_count += 1
                    except Exception as e:
                        error_rows.append(f"第{i}行: {str(e)}")
            else:
                wb = load_workbook(file)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                for i, row in enumerate(rows, start=2):
                    try:
                        data = {}
                        for j, col in enumerate(config['columns']):
                            if j < len(row):
                                val = row[j]
                                if config['types'][j] == 'int':
                                    data[col] = int(val) if val else None
                                elif config['types'][j] == 'float':
                                    data[col] = float(val) if val else 0
                                else:
                                    data[col] = str(val) if val else ''
                        if 'status' not in data:
                            data['status'] = 1
                        db.execute(f"INSERT INTO {table} ({','.join(config['columns'])}) VALUES ({','.join(['?']*len(config['columns']))})",
                                   [data.get(c) for c in config['columns']])
                        success_count += 1
                    except Exception as e:
                        error_rows.append(f"第{i}行: {str(e)}")

            if error_rows:
                db.rollback()
                return jsonify({
                    'code': 400,
                    'message': '导入失败，文件中的数据未写入',
                    'data': {
                        'success': 0,
                        'errors': error_rows[:10],
                    },
                })

            db.commit()
            return jsonify({
                'code': 0,
                'message': f'导入成功 {success_count} 条',
                'data': {'success': success_count, 'errors': error_rows[:10]}
            })
        except Exception as e:
            db.rollback()
            return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'})

    @app.route('/api/template/<table>')
    @login_required
    def download_template(table):
        if table not in TABLE_CONFIG:
            return jsonify({'code': 400, 'message': '不支持的表'})
        config = TABLE_CONFIG[table]
        if config.get('importable') is False:
            return jsonify({'code': 400, 'message': '不支持导入该表'})

        wb = Workbook()
        ws = wb.active
        ws.title = config['name']
        ws.append(config['headers'])

        example = []
        for t in config['types']:
            if t == 'int':
                example.append(1)
            elif t == 'float':
                example.append(100.0)
            else:
                example.append('示例')
        ws.append(example)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        filename = f"{table}_template.xlsx"
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    # 物料追溯 API
    @app.route('/api/trace/batch/list')
    @login_required
    def trace_batch_list():
        db = get_db()
        try:
            page = max(1, int(request.args.get('page', 1)))
            size = min(500, max(1, int(request.args.get('size', 15))))
        except (TypeError, ValueError):
            return jsonify({'code': 400, 'message': '分页参数必须是整数'}), 400
        keyword = request.args.get('keyword', '')
        if keyword:
            total = db.execute("SELECT COUNT(*) as c FROM inv_batch WHERE batch_no LIKE ?", (f'%{keyword}%',)).fetchone()['c']
            rows = db.execute('''SELECT b.*, p.product_name, p.code as product_code
                FROM inv_batch b LEFT JOIN base_product p ON b.product_id=p.id
                WHERE b.batch_no LIKE ? ORDER BY b.id DESC LIMIT ? OFFSET ?''',
                (f'%{keyword}%', size, (page-1)*size)).fetchall()
        else:
            total = db.execute("SELECT COUNT(*) as c FROM inv_batch").fetchone()['c']
            rows = db.execute('''SELECT b.*, p.product_name, p.code as product_code
                FROM inv_batch b LEFT JOIN base_product p ON b.product_id=p.id
                ORDER BY b.id DESC LIMIT ? OFFSET ?''', (size, (page-1)*size)).fetchall()
        return jsonify({'code': 0, 'data': {'list': [dict(r) for r in rows], 'total': total}})

    @app.route('/api/trace/batch/add', methods=['POST'])
    @login_required
    def trace_batch_add():
        d = request.get_json(silent=True)
        if not isinstance(d, dict):
            return jsonify({'code': 400, 'message': '请求数据必须是JSON对象'}), 400
        batch_no = str(d.get('batch_no') or '').strip()
        try:
            product_id = int(d.get('product_id'))
            quantity = float(d.get('quantity') or 0)
        except (TypeError, ValueError):
            return jsonify({'code': 400, 'message': '产品或数量格式错误'}), 400
        if not batch_no:
            return jsonify({'code': 400, 'message': '批次号不能为空'}), 400
        if quantity <= 0:
            return jsonify({'code': 400, 'message': '批次数量必须大于0'}), 400
        db = get_db()
        if not db.execute('SELECT 1 FROM base_product WHERE id=?', (product_id,)).fetchone():
            return jsonify({'code': 404, 'message': '产品不存在'}), 404
        try:
            cursor = db.execute("INSERT INTO inv_batch (batch_no,product_id,supplier,quantity,production_date,expiry_date,status) VALUES (?,?,?,?,?,?,?)",
                       (batch_no, product_id, str(d.get('supplier') or '').strip(), quantity,
                        d.get('production_date',''), d.get('expiry_date',''), d.get('status',1)))
            db.commit()
            return jsonify({'code': 0, 'message': '批次新增成功', 'data': {'id': cursor.lastrowid}})
        except INTEGRITY_ERRORS:
            db.rollback()
            return jsonify({'code': 409, 'message': '批次号已存在'}), 409

    @app.route('/api/trace/batch/delete', methods=['POST'])
    @login_required
    def trace_batch_delete():
        d = request.get_json(silent=True) or {}
        if not d.get('id'):
            return jsonify({'code': 400, 'message': '缺少批次ID'}), 400
        db = get_db()
        if db.execute('SELECT 1 FROM inv_trace WHERE batch_id=? LIMIT 1', (d['id'],)).fetchone():
            return jsonify({'code': 409, 'message': '批次已有追溯记录，不能删除'}), 409
        cursor = db.execute("DELETE FROM inv_batch WHERE id=?", (d['id'],))
        if cursor.rowcount == 0:
            db.rollback()
            return jsonify({'code': 404, 'message': '批次不存在'}), 404
        db.commit()
        return jsonify({'code': 0, 'message': '删除成功'})

    @app.route('/api/trace/chain/<int:batch_id>')
    @login_required
    def trace_chain(batch_id):
        db = get_db()
        batch = db.execute('''SELECT b.*, p.product_name FROM inv_batch b
            LEFT JOIN base_product p ON b.product_id=p.id WHERE b.id=?''', (batch_id,)).fetchone()
        if not batch:
            return jsonify({'code': 404, 'message': '批次不存在'}), 404
        traces = db.execute("SELECT * FROM inv_trace WHERE batch_id=? ORDER BY created_at", (batch_id,)).fetchall()
        return jsonify({'code': 0, 'data': {'batch': dict(batch), 'traces': [dict(r) for r in traces]}})

    @app.route('/api/trace/add', methods=['POST'])
    @login_required
    def trace_add():
        d = request.get_json(silent=True)
        if not isinstance(d, dict):
            return jsonify({'code': 400, 'message': '请求数据必须是JSON对象'}), 400
        from flask import session
        db = get_db()
        try:
            batch_id = int(d.get('batch_id'))
            quantity = float(d.get('quantity') or 0)
        except (TypeError, ValueError):
            return jsonify({'code': 400, 'message': '批次或数量格式错误'}), 400
        trace_type = str(d.get('trace_type') or '').strip()
        if not trace_type:
            return jsonify({'code': 400, 'message': '追溯类型不能为空'}), 400
        if not db.execute('SELECT 1 FROM inv_batch WHERE id=?', (batch_id,)).fetchone():
            return jsonify({'code': 404, 'message': '批次不存在'}), 404
        ref_no = str(d.get('ref_no') or '').strip()
        cursor = db.execute("INSERT INTO inv_trace (batch_id,trace_type,biz_no,operation,ref_no,ref_id,quantity,operator,remark) VALUES (?,?,?,?,?,?,?,?,?)",
                   (batch_id, trace_type, ref_no, trace_type, ref_no, d.get('ref_id',0),
                    quantity, session.get('user_id'), d.get('remark','')))
        db.commit()
        return jsonify({'code': 0, 'message': '追溯记录新增成功', 'data': {'id': cursor.lastrowid}})

    @app.route('/api/trace/query')
    @login_required
    def trace_query():
        db = get_db()
        keyword = request.args.get('keyword', '')
        if not keyword:
            return jsonify({'code': 0, 'data': []})
        batches = db.execute('''SELECT b.*, p.product_name FROM inv_batch b
            LEFT JOIN base_product p ON b.product_id=p.id
            WHERE b.batch_no LIKE ? OR p.product_name LIKE ?
            ORDER BY b.id DESC LIMIT 50''', (f'%{keyword}%', f'%{keyword}%')).fetchall()
        results = []
        for b in batches:
            traces = db.execute("SELECT * FROM inv_trace WHERE batch_id=? ORDER BY created_at", (b['id'],)).fetchall()
            results.append({'batch': dict(b), 'traces': [dict(t) for t in traces]})
        return jsonify({'code': 0, 'data': results})

    return app


app = create_app()

if __name__ == '__main__':
    init_db()
    _init_extra_tables()
    from machine_runtime import MachineCommunicationRuntime
    machine_runtime = MachineCommunicationRuntime()
    machine_runtime.start()
    print("=" * 50)
    print("  MES工厂管家 启动成功!")
    print("  访问地址: http://localhost:8080")
    print("  默认账号: admin / admin123")
    print("=" * 50)
    try:
        host = os.environ.get('MES_HOST', '0.0.0.0')
        port = int(os.environ.get('MES_PORT', '8080'))
        workers = max(1, int(os.environ.get('MES_WORKERS', '4')))
        production_mode = os.environ.get('MES_ENV', '').lower() == 'production'
        if production_mode or os.environ.get('MES_SERVER', '').lower() == 'waitress':
            try:
                from waitress import serve
            except ImportError as exc:
                raise RuntimeError('Waitress is required for production mode') from exc
            serve(app, host=host, port=port, threads=workers * 2)
        else:
            app.run(host=host, port=port, debug=False)
    finally:
        machine_runtime.stop()
